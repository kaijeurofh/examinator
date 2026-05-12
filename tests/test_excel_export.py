"""Tests for `examinator.core.excel_export`."""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from examinator.core.excel_export import filename_for, to_xlsx_bytes
from examinator.core.schemas import (
    AcademicLevel,
    DifficultyLevel,
    HausarbeitQAPair,
    KlausurQAPair,
    PageQuestions,
    TaskType,
)


def _hausarbeit_pair(i: int) -> HausarbeitQAPair:
    return HausarbeitQAPair(
        question=f"Frage Nr. {i}?",
        question_type="Analytische Frage",
        academic_level=AcademicLevel.BACHELOR,
        scope="3.500-4.500 Woerter",
        core_topic="Marketing",
        guideline_examiner="Examiner notes.",
        guideline_student="Student roadmap.",
        bewertungsschema_rubric="- Formale: 24\n- Inhalt: 96",
        source_page=i,
    )


def _klausur_pair(i: int) -> KlausurQAPair:
    return KlausurQAPair(
        question=f"Klausurfrage {i}?",
        klausur_subtype="Knowledge",
        difficulty_level=DifficultyLevel.MITTEL,
        academic_level=AcademicLevel.MASTER,
        scope="20 Punkte",
        core_topic="Statistik",
        guideline_examiner="Pruefer Hinweise.",
        guideline_student="Studierenden Hinweise.",
        musterloesung_text="Musterloesungstext " * 30,
        musterloesung_rubric="Formal 6, Inhalt 14.",
        source_page=i,
    )


def test_export_hausarbeit_to_xlsx_round_trips() -> None:
    container: PageQuestions[HausarbeitQAPair] = PageQuestions(
        questions=[_hausarbeit_pair(i) for i in range(1, 11)]
    )
    blob = to_xlsx_bytes(container, TaskType.HAUSARBEIT)
    assert blob[:2] == b"PK"  # .xlsx is a zip

    wb = load_workbook(BytesIO(blob))
    assert wb.active is not None
    sheet = wb.active
    assert sheet.title == "Hausarbeitsfragen"
    # Header row + 10 data rows.
    assert sheet.max_row == 11
    assert sheet["A1"].value == "Nr."
    assert sheet["B1"].value == "Frage"
    assert sheet["A2"].value == 1
    assert sheet["B2"].value == "Frage Nr. 1?"
    assert sheet["J2"].value == 1  # source_page column


def test_export_klausur_includes_musterloesung_columns() -> None:
    container: PageQuestions[KlausurQAPair] = PageQuestions(
        questions=[_klausur_pair(i) for i in range(1, 4)]
    )
    blob = to_xlsx_bytes(container, TaskType.KLAUSUR)
    wb = load_workbook(BytesIO(blob))
    assert wb.active is not None
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]
    assert "Musterloesung" in headers
    assert "Bewertungsrubrik" in headers


def test_filename_for_each_task_type_is_distinct() -> None:
    names = {filename_for(t) for t in TaskType}
    assert len(names) == 4
    for name in names:
        assert name.endswith(".xlsx")
