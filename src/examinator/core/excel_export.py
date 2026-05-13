"""Render a `PageQuestions[T]` result to an `.xlsx` workbook.

Each task type has its own column layout — we keep one sheet per workbook
("Fragen") and set sensible column widths plus wrap-text on the long-form
cells (guidelines, rubrics, model answers) so the file is readable in Excel
right after download.
"""

from __future__ import annotations

from collections.abc import Iterable
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from examinator.core.schemas import (
    PageQuestions,
    QAPair,
    TaskType,
)

_HEADER_FILL = PatternFill(start_color="FF1F3A5F", end_color="FF1F3A5F", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_WRAP = Alignment(wrap_text=True, vertical="top")
_TOP = Alignment(vertical="top")

# Column spec: (header, attr_name, width, wrap).
_ColumnSpec = tuple[str, str, int, bool]

_HAUSARBEIT_COLS: tuple[_ColumnSpec, ...] = (
    ("Nr.", "_index", 6, False),
    ("Frage", "question", 60, True),
    ("Fragetyp", "question_type", 28, True),
    ("Niveau", "academic_level", 12, False),
    ("Umfang", "scope", 22, True),
    ("Kernthema", "core_topic", 30, True),
    ("Hinweise Pruefende", "guideline_examiner", 50, True),
    ("Hinweise Studierende", "guideline_student", 70, True),
    ("Bewertungsschema", "bewertungsschema_rubric", 70, True),
    ("Seite", "source_page", 8, False),
)

_PROJEKTARBEIT_COLS: tuple[_ColumnSpec, ...] = (
    ("Nr.", "_index", 6, False),
    ("Frage", "question", 60, True),
    ("Fragetyp", "question_type", 18, False),
    ("Niveau", "academic_level", 12, False),
    ("Ausfuehrungsformat", "execution_format", 32, True),
    ("Umfang", "scope", 22, True),
    ("Kernthema", "core_topic", 30, True),
    ("Hinweise Pruefende", "guideline_examiner", 50, True),
    ("Hinweise Studierende", "guideline_student", 70, True),
    ("Bewertungsschema", "bewertungsschema_rubric", 70, True),
    ("Seite", "source_page", 8, False),
)

_KLAUSUR_COLS: tuple[_ColumnSpec, ...] = (
    ("Nr.", "_index", 6, False),
    ("Frage", "question", 60, True),
    ("Subtyp", "klausur_subtype", 22, False),
    ("Schwierigkeit", "difficulty_level", 14, False),
    ("Niveau", "academic_level", 12, False),
    ("Umfang", "scope", 22, True),
    ("Kernthema", "core_topic", 30, True),
    ("Hinweise Pruefende", "guideline_examiner", 50, True),
    ("Hinweise Studierende", "guideline_student", 50, True),
    ("Musterloesung", "musterloesung_text", 80, True),
    ("Bewertungsrubrik", "musterloesung_rubric", 60, True),
    ("Seite", "source_page", 8, False),
)

_EINSENDE_COLS: tuple[_ColumnSpec, ...] = (
    ("Nr.", "_index", 6, False),
    ("Frage", "question", 60, True),
    ("Subtyp", "einsende_subtype", 22, False),
    ("Schwierigkeit", "difficulty_level", 14, False),
    ("Niveau", "academic_level", 12, False),
    ("Umfang", "scope", 22, True),
    ("Kernthema", "core_topic", 30, True),
    ("Hinweise Pruefende", "guideline_examiner", 50, True),
    ("Hinweise Studierende", "guideline_student", 50, True),
    ("Musterloesung", "musterloesung_text", 80, True),
    ("Bewertungsrubrik", "musterloesung_rubric", 60, True),
    ("Seite", "source_page", 8, False),
)

_COLS_FOR: dict[TaskType, tuple[_ColumnSpec, ...]] = {
    TaskType.HAUSARBEIT: _HAUSARBEIT_COLS,
    TaskType.PROJEKTARBEIT: _PROJEKTARBEIT_COLS,
    TaskType.KLAUSUR: _KLAUSUR_COLS,
    TaskType.EINSENDEAUFGABE: _EINSENDE_COLS,
}

_SHEET_TITLE: dict[TaskType, str] = {
    TaskType.HAUSARBEIT: "Hausarbeitsfragen",
    TaskType.PROJEKTARBEIT: "Projektarbeitsfragen",
    TaskType.KLAUSUR: "Klausurfragen",
    TaskType.EINSENDEAUFGABE: "Einsendeaufgaben",
}


def to_xlsx_bytes(result: PageQuestions, task_type: TaskType) -> bytes:  # type: ignore[type-arg]
    """Render a result to an in-memory `.xlsx` byte blob."""
    workbook = Workbook()
    sheet = workbook.active
    if sheet is None:  # pragma: no cover - openpyxl always creates one
        sheet = workbook.create_sheet()
    sheet.title = _SHEET_TITLE[task_type]

    columns = _COLS_FOR[task_type]
    _write_header(sheet, columns)
    _write_rows(sheet, columns, result.questions)
    _apply_column_widths(sheet, columns)
    sheet.freeze_panes = "A2"

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def filename_for(task_type: TaskType) -> str:
    """Return a download filename consistent with the sheet title."""
    return f"examinator_{task_type.value}.xlsx"


def _write_header(sheet: Worksheet, columns: tuple[_ColumnSpec, ...]) -> None:
    for col_idx, (header, _, _, _) in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _WRAP
    sheet.row_dimensions[1].height = 28


def _write_rows(
    sheet: Worksheet,
    columns: tuple[_ColumnSpec, ...],
    questions: Iterable[QAPair],
) -> None:
    for row_offset, question in enumerate(questions, start=2):
        row_index = row_offset - 1
        for col_idx, (_, attr, _, wrap) in enumerate(columns, start=1):
            value: object = row_index if attr == "_index" else _coerce(getattr(question, attr, ""))
            cell = sheet.cell(row=row_offset, column=col_idx, value=value)
            cell.alignment = _WRAP if wrap else _TOP


def _apply_column_widths(sheet: Worksheet, columns: tuple[_ColumnSpec, ...]) -> None:
    for col_idx, (_, _, width, _) in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def _coerce(value: object) -> str | int | float:
    """Make pydantic / enum values openpyxl-friendly."""
    if value is None:
        return ""
    if isinstance(value, str | int | float):
        return value
    # StrEnum stringifies via str(); fallthrough handles everything else.
    return str(value)
